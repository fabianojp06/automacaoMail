from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo = "E:\\projRPA\\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#seleciona a planilha com todos os dados
sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = "" 
totalLInha = len(sheet_selecionada['A']) + 1

for linha in range(2,len(sheet_selecionada['A']) + 1):

    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:

        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A'])+ 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)

        #preenche os dados
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)



    else:

        #adiciona o nome do funcionario que esta na linha que o cpodigo está passando
        nomeNovo = sheet_selecionada['A%s' % linha].value

        novaPlanilha = criandoNovoArquivoExcel.active

        novaPlanilha.title = "Vendas"

        caminhoNovaPlanilha = "E:\\projRPA\\" + "sheet_selecionada['A%s' % linha].value" +".xlsx"


        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        #inserindo os nomes das colunas
        selecionaSheetVendasNovaPlanilha['A1'] = 'Vendedor'
        selecionaSheetVendasNovaPlanilha['B1'] = 'Produtos'
        selecionaSheetVendasNovaPlanilha['C1'] = 'Vendas'

        #preenche as informaçoes da segunda linha
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheetVendasNovaPlanilha.delete_rows(3,100)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

#salva as alterações feitas na planilha
planilha_aberta.save(filename=nome_arquivo)

#abre o arquivo com as alterações realizadas
#os.startfile(nome_arquivo)

print('Planilhas criadas com sucesso!')






